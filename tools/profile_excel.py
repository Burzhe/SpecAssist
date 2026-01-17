from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook

from config import EXCEL_PATH

SAMPLE_ROWS = 2000
TOP_EXAMPLES = 20

DIMENSION_REGEX = re.compile(
    r"(\d+(?:[.,]\d+)?)\s*(?:x|×|\*)\s*(\d+(?:[.,]\d+)?)(?:\s*(?:x|×|\*)\s*(\d+(?:[.,]\d+)?))?\s*(?:мм|mm|см|cm)?",
    re.IGNORECASE,
)
IMAGE_EXT_REGEX = re.compile(r"\.(jpg|jpeg|png|gif|webp|bmp|tiff)(\?.*)?$", re.IGNORECASE)

PRICE_TOKENS = ("price", "cost", "amount", "цена", "стоим", "руб")
DESCRIPTION_TOKENS = ("desc", "опис", "name", "title", "comment", "примеч")
TYPE_TOKENS = ("type", "тип", "катег", "class", "вид")
DIMENSION_TOKENS = (
    "size",
    "dimension",
    "размер",
    "габарит",
    "width",
    "height",
    "depth",
    "length",
    "шир",
    "выс",
    "глуб",
    "длин",
)
LIGHT_TOKENS = ("light", "подсвет", "led", "свет", "лента", "illum")
PHOTO_TOKENS = ("photo", "image", "img", "картин", "фото", "изобр")
IDENTIFIER_TOKENS = (
    "id",
    "sku",
    "арт",
    "артикул",
    "article",
    "код",
    "uid",
    "uuid",
    "barcode",
    "штрих",
    "source",
)


@dataclass
class SheetProfile:
    name: str
    rows: int
    columns: int


def _contains_any(text: str, tokens: Iterable[str]) -> bool:
    lower = text.lower()
    return any(token in lower for token in tokens)


def _sample_text_values(series: pd.Series, limit: int = 200) -> list[str]:
    return (
        series.dropna()
        .astype(str)
        .map(str.strip)
        .loc[lambda s: s != ""]
        .head(limit)
        .tolist()
    )


def _detect_dimensions(values: Iterable[str]) -> bool:
    return any(DIMENSION_REGEX.search(value) for value in values)


def _detect_light(values: Iterable[str]) -> bool:
    return any(_contains_any(value, LIGHT_TOKENS) for value in values)


def _detect_photo(values: Iterable[str]) -> bool:
    for value in values:
        if IMAGE_EXT_REGEX.search(value):
            return True
        if value.startswith("http") and _contains_any(value, PHOTO_TOKENS):
            return True
    return False


def _top_examples(series: pd.Series) -> list[dict[str, Any]]:
    if series.empty:
        return []
    cleaned = series.dropna().astype(str).map(str.strip)
    cleaned = cleaned.loc[cleaned != ""]
    if cleaned.empty:
        return []
    counts = Counter(cleaned)
    most_common = counts.most_common(TOP_EXAMPLES)
    return [{"value": value, "count": count} for value, count in most_common]


def _profile_column(name: str, series: pd.Series) -> dict[str, Any]:
    total = len(series)
    null_percent = float(series.isna().mean() * 100) if total else 0.0
    sample_values = _sample_text_values(series)
    lower_name = name.lower()

    is_price = _contains_any(lower_name, PRICE_TOKENS)
    is_description = _contains_any(lower_name, DESCRIPTION_TOKENS)
    is_type = _contains_any(lower_name, TYPE_TOKENS)
    is_dimensions = _contains_any(lower_name, DIMENSION_TOKENS) or _detect_dimensions(sample_values)
    is_light = _contains_any(lower_name, LIGHT_TOKENS) or _detect_light(sample_values)
    is_photo = _contains_any(lower_name, PHOTO_TOKENS) or _detect_photo(sample_values)
    is_identifier = _contains_any(lower_name, IDENTIFIER_TOKENS)

    if not is_price and pd.api.types.is_numeric_dtype(series) and _contains_any(lower_name, ("price", "цена")):
        is_price = True

    return {
        "name": name,
        "dtype": str(series.dtype),
        "null_percent": round(null_percent, 2),
        "top_values": _top_examples(series),
        "candidates": {
            "price": is_price,
            "description": is_description,
            "type": is_type,
            "dimensions": is_dimensions,
            "light": is_light,
            "photo": is_photo,
            "source_identifier": is_identifier,
        },
    }


def _load_sheet_profiles(excel_path: Path) -> list[SheetProfile]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    profiles: list[SheetProfile] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        profiles.append(
            SheetProfile(name=sheet_name, rows=sheet.max_row or 0, columns=sheet.max_column or 0)
        )
    workbook.close()
    return profiles


def _profile_sheet(excel_path: Path, sheet: SheetProfile) -> dict[str, Any]:
    df = pd.read_excel(
        excel_path,
        sheet_name=sheet.name,
        nrows=SAMPLE_ROWS,
        engine="openpyxl",
    )

    columns_profile = [_profile_column(name, df[name]) for name in df.columns]
    candidate_columns: dict[str, list[str]] = {
        "price": [],
        "description": [],
        "type": [],
        "dimensions": [],
        "light": [],
        "photo": [],
        "source_identifier": [],
    }
    for column in columns_profile:
        for key, enabled in column["candidates"].items():
            if enabled:
                candidate_columns[key].append(column["name"])

    return {
        "name": sheet.name,
        "rows": sheet.rows,
        "columns": sheet.columns,
        "sample_rows": int(df.shape[0]),
        "column_names": list(df.columns),
        "columns_profile": columns_profile,
        "candidate_columns": candidate_columns,
    }


def _write_outputs(output_dir: Path, data: dict[str, Any]) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "excel_profile.json"
    md_path = output_dir / "excel_profile.md"

    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = ["# Excel profile", "", f"Source: {data['excel_path']}", ""]
    for sheet in data["sheets"]:
        lines.extend(
            [
                f"## Sheet: {sheet['name']}",
                "",
                f"Rows (estimated): {sheet['rows']}",
                f"Columns: {sheet['columns']}",
                f"Sample rows: {sheet['sample_rows']}",
                "",
                "### Candidate columns",
            ]
        )
        for key, values in sheet["candidate_columns"].items():
            label = key.replace("_", " ").title()
            if values:
                lines.append(f"- {label}: {', '.join(values)}")
            else:
                lines.append(f"- {label}: (none)")
        lines.append("")
        lines.append("### Columns")
        for column in sheet["columns_profile"]:
            lines.append(f"- **{column['name']}** ({column['dtype']})")
            lines.append(f"  - Null %: {column['null_percent']}")
            if column["top_values"]:
                examples = ", ".join(
                    f"{item['value']} ({item['count']})" for item in column["top_values"]
                )
                lines.append(f"  - Top values: {examples}")
            else:
                lines.append("  - Top values: (none)")
        lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")
    return json_path, md_path


def _print_summary(data: dict[str, Any]) -> None:
    print(f"Excel profile for {data['excel_path']}")
    print(f"Sheets: {len(data['sheets'])}")
    for sheet in data["sheets"]:
        print(f"- {sheet['name']}: {sheet['rows']} rows, {sheet['columns']} cols")
        for key, values in sheet["candidate_columns"].items():
            if values:
                print(f"  {key}: {', '.join(values)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Profile Excel sheets")
    parser.add_argument("--excel", type=str, default=str(EXCEL_PATH), help="Path to Excel file")
    parser.add_argument(
        "--output-dir",
        type=str,
        default="data",
        help="Directory for excel_profile.json and excel_profile.md",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser()
    sheet_profiles = _load_sheet_profiles(excel_path)
    sheets = [_profile_sheet(excel_path, sheet) for sheet in sheet_profiles]

    data = {"excel_path": str(excel_path), "sheets": sheets}
    json_path, md_path = _write_outputs(Path(args.output_dir), data)
    _print_summary(data)
    print(f"Saved: {json_path}")
    print(f"Saved: {md_path}")


if __name__ == "__main__":
    main()
